# -*- coding: utf-8 -*-
import datetime
import ast
import os
import sys
import math
import calendar
import time
import datetime as dt
from datetime import timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from copy import copy

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D

import warnings
import importlib.util

warnings.filterwarnings('ignore')

def _hours_module_candidates() -> List[Path]:
    candidates: List[Path] = []
    try:
        base = Path(__file__).resolve()
        candidates.append(base.parents[1] / "Проверка чеков" / "Hours_from_checks_simple.py")
        candidates.append(base.parent / "Hours_from_checks_simple.py")
    except Exception:
        pass
    cwd = Path.cwd()
    candidates.append(cwd / "Проверка чеков" / "Hours_from_checks_simple.py")
    candidates.append(cwd / "Hours_from_checks_simple.py")
    uniq: List[Path] = []
    for item in candidates:
        if item not in uniq:
            uniq.append(item)
    return uniq


def load_hours_calculator() -> Tuple[Optional[type], Optional[Exception], Optional[Path]]:
    last_error: Optional[Exception] = None
    for candidate in _hours_module_candidates():
        if not candidate.exists():
            continue
        try:
            spec = importlib.util.spec_from_file_location("hours_from_checks_simple", candidate)
            if spec and spec.loader:
                module = importlib.util.module_from_spec(spec)
                sys.modules[spec.name] = module
                spec.loader.exec_module(module)
                hours_cls = getattr(module, "HoursCalculator", None)
                if hours_cls is not None:
                    return hours_cls, None, candidate
        except Exception as exc:
            last_error = exc
    return None, last_error, None


HoursCalculator, _hours_import_error, _hours_import_path = load_hours_calculator()

from constants import (SRC_PATH, PATH_TO_DAY_NIGHT_FILE, LAST_FILE_FROM_PATH_WEIGHTS,
                       LAST_FILE_FROM_PATH_CURRENT, OUTPUT_MODELS_COMPARISON,
                       NAME_OF_COLUMN_OF_KSSS, NAME_OF_COLUMN_OF_DATE,
                       NAME_OF_COLUMN_OF_SEGMENT, HORIZON, DAYS_UNTIL_SENDINGS, RESULT_PREDICTIONS,
                       MONTH_NAMES, AZS_PRED_FILE, SEASONAL_PERIOD, WEEKDAY)
from constant_functions import last_date_indicator
from current_forecast import periods_maker, indicator_adder, proportion_of_day_night

# =====================
# Load inputs
# =====================

# =====================
# Target month name
beginning_time = time.time()
in_work = 0
AZS_for_pred = pd.read_excel(AZS_PRED_FILE)['КССС'].tolist()
# Optional overrides from platform GUI
CHECKS_PATH_OVERRIDE = os.environ.get("CHECKS_PATH_OVERRIDE")
FORECAST_START = os.environ.get("FORECAST_START")
FORECAST_END = os.environ.get("FORECAST_END")
CHECKS_DELIMITER = os.environ.get("CHECKS_DELIMITER")
AZS_LIMIT_COUNT = os.environ.get("AZS_LIMIT_COUNT")
KSSS_RANGE_START = os.environ.get("KSSS_RANGE_START")
KSSS_RANGE_END = os.environ.get("KSSS_RANGE_END")
KSSS_LIST = os.environ.get("KSSS_LIST")
PROGRESS_FILE = os.environ.get("PROGRESS_FILE")
STOP_FILE = os.environ.get("STOP_FILE")
RESULT_DIR_FILE = os.environ.get("RESULT_DIR_FILE")
# =====================

progress_path = Path(PROGRESS_FILE) if PROGRESS_FILE else None
stop_path = Path(STOP_FILE) if STOP_FILE else None
result_dir_path = Path(RESULT_DIR_FILE) if RESULT_DIR_FILE else None


def write_progress(done: int, total: int) -> None:
    if progress_path is None:
        return
    try:
        progress_path.parent.mkdir(parents=True, exist_ok=True)
        with progress_path.open("w", encoding="utf-8") as f:
            f.write(f"{done}/{total}")
    except Exception:
        pass


def stop_requested() -> bool:
    try:
        return stop_path is not None and stop_path.exists()
    except Exception:
        return False


def write_result_dir(path: str | Path) -> None:
    if result_dir_path is None:
        return
    try:
        result_dir_path.parent.mkdir(parents=True, exist_ok=True)
        result_dir_path.write_text(str(path), encoding="utf-8")
    except Exception:
        pass

def normalize_ksss(value) -> str:
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def parse_delimiter(value: Optional[str]) -> str:
    if not value:
        return "|"
    v = str(value).strip()
    if v.lower() in ("\\t", "tab", "tabs", "t"):
        return "\t"
    return v

def parse_ksss_list(value: Optional[str]) -> List[str]:
    if not value:
        return []
    parts = []
    for item in str(value).split(","):
        item = item.strip()
        if item:
            s = str(item).strip()
            if s.endswith(".0"):
                s = s[:-2]
            parts.append(s)
    return parts

def coerce_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None

checks_sep = parse_delimiter(CHECKS_DELIMITER)
checks_source_path = Path(CHECKS_PATH_OVERRIDE) if CHECKS_PATH_OVERRIDE else Path(SRC_PATH)
checks = pd.read_csv(checks_source_path, sep=checks_sep)
checks_len_raw = len(checks)
day_night = pd.read_csv(PATH_TO_DAY_NIGHT_FILE, sep='|')


AZS_for_pred = [normalize_ksss(x) for x in AZS_for_pred]
selected_ksss = AZS_for_pred
ksss_list_env = parse_ksss_list(KSSS_LIST)
if ksss_list_env:
    selected_ksss = [k for k in selected_ksss if k in set(ksss_list_env)]

range_start = coerce_int(KSSS_RANGE_START)
range_end = coerce_int(KSSS_RANGE_END)
if range_start is not None and range_end is not None:
    if range_start > range_end:
        range_start, range_end = range_end, range_start
    range_filtered = []
    for k in selected_ksss:
        k_int = coerce_int(k)
        if k_int is not None and range_start <= k_int <= range_end:
            range_filtered.append(k)
    selected_ksss = range_filtered

limit_count = coerce_int(AZS_LIMIT_COUNT)
if limit_count is not None and limit_count > 0:
    selected_ksss = selected_ksss[:limit_count]

checks = checks[checks[NAME_OF_COLUMN_OF_KSSS].astype(str).str.strip().isin(selected_ksss)]
checks_len_azs = len(checks)
sp_final = [
    i
    for i in checks[NAME_OF_COLUMN_OF_KSSS].unique()
    if len(checks[checks[NAME_OF_COLUMN_OF_KSSS] == i]) >= 365
]
checks = checks[checks[NAME_OF_COLUMN_OF_KSSS].isin(sp_final)]
checks_len_min_days = len(checks)


def year_len(year):
    days_in_year = 365
    if calendar.monthrange(year=year, month=2)[1] == 29:
        days_in_year = 366
    return days_in_year

def month_len(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]

# Standardize column names
def clean_cols(df):
    df.columns = (
        df.columns.str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    return df

def parse_env_date(value: Optional[str]) -> Optional[pd.Timestamp]:
    if not value:
        return None
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return pd.to_datetime(value, format=fmt)
        except Exception:
            continue
    try:
        return pd.to_datetime(value)
    except Exception:
        return None

checks = clean_cols(checks)
checks['date'] = pd.to_datetime(checks[['year','month','day']], errors='coerce')
checks = checks[checks[NAME_OF_COLUMN_OF_DATE] <= last_date_indicator(checks_source_path, sep=checks_sep)]
checks_len_final = len(checks)

if checks.empty:
    print("Нет данных для прогноза после фильтров.")
    print(f"Всего строк: {checks_len_raw}")
    print(f"После фильтра АЗС: {checks_len_azs}")
    print(f"После фильтра 365 дней: {checks_len_min_days}")
    print(f"После фильтра даты: {checks_len_final}")
    write_progress(0, 0)
    sys.exit(0)

day_night = clean_cols(day_night)
day_night['date'] = pd.to_datetime(day_night[['year','month','day']], errors='coerce')

KSSS_COL = DN_KSSS_COL = NAME_OF_COLUMN_OF_KSSS
DATE_COL = DN_VALUE_COL = NAME_OF_COLUMN_OF_DATE
VALUE_COL = 'unique_cheques'

# =====================
# Input readers for MAPE sources
# =====================
def coerce_numeric(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    s = str(x).strip().replace("%", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan

def parse_weight_tuple(s):
    if pd.isna(s):
        return (np.nan, np.nan, np.nan)
    if isinstance(s, (tuple, list)) and len(s) == 3:
        return tuple(float(v) for v in s)
    s = str(s).strip().replace(",", ", ")
    try:
        tup = ast.literal_eval(s)
        if isinstance(tup, tuple) and len(tup) == 3:
            return (float(tup[0]), float(tup[1]), float(tup[2]))
    except Exception:
        nums = []
        cur = ""
        for ch in s:
            if ch in "0123456789.-":
                cur += ch
            else:
                if cur:
                    nums.append(cur)
                    cur = ""
        if cur:
            nums.append(cur)
        if len(nums) >= 3:
            try:
                return (float(nums[0]), float(nums[1]), float(nums[2]))
            except Exception:
                pass
    return (np.nan, np.nan, np.nan)

def read_weights_csv(path: str) -> pd.DataFrame:
    df = None
    last_err = None
    for enc in ("utf-8-sig", "cp1251"):
        try:
            df = pd.read_csv(path, encoding=enc, sep=None, engine="python")
            break
        except Exception as e:
            last_err = e
    if df is None:
        raise last_err
    return clean_cols(df)

def read_current_mape_csv(path: str) -> pd.DataFrame:
    df = None
    last_err = None
    for enc in ("utf-8-sig", "cp1251"):
        try:
            df = pd.read_csv(path, encoding=enc, sep=None, engine="python")
            break
        except Exception as e:
            last_err = e
    if df is None:
        raise last_err

    df = clean_cols(df)
    key_col_candidates = [
        c for c in df.columns
        if "КССС|MAPE" in c or "KCCC|MAPE" in c or "KSSS|MAPE" in c
    ]
    if not key_col_candidates:
        raise ValueError("Не найден столбец вида 'КССС|MAPE' в файле current.")
    key_col = key_col_candidates[0]

    split = df[key_col].astype(str).str.split("|", n=1, expand=True)
    df["KSSS"] = split[0].str.strip()
    df["current_mape_pct"] = split[1].apply(coerce_numeric)
    return df[["KSSS", "current_mape_pct"]]

weights_raw = read_weights_csv(LAST_FILE_FROM_PATH_WEIGHTS)

ksss_col_candidates = [c for c in weights_raw.columns if "КССС" in c.upper() or "KSSS" in c.upper()]
weights_col_candidates = [c for c in weights_raw.columns if "ВЕС" in c.upper()]
mape_col_candidates = [c for c in weights_raw.columns if "МАРЕ" in c.upper() or "MAPE" in c.upper()]

if not ksss_col_candidates or not weights_col_candidates or not mape_col_candidates:
    raise ValueError("Не удалось определить колонки КССС/Веса/МАРЕ в файле весов.")

weights_df = weights_raw.rename(columns={
    ksss_col_candidates[0]: "KSSS",
    weights_col_candidates[0]: "Веса",
    mape_col_candidates[0]: "weights_mape_pct"
})

weights_df["KSSS"] = weights_df["KSSS"].apply(normalize_ksss)
weights_df[["w1", "w2", "w3"]] = weights_df["Веса"].apply(parse_weight_tuple).apply(pd.Series)
weights_df["weights_mape_pct"] = weights_df["weights_mape_pct"].apply(coerce_numeric)

current_mape_df = read_current_mape_csv(LAST_FILE_FROM_PATH_CURRENT)
current_mape_df["KSSS"] = current_mape_df["KSSS"].apply(normalize_ksss)

weights_map = (
    weights_df.set_index("KSSS")[["w1", "w2", "w3", "weights_mape_pct"]].to_dict("index")
)
current_mape_map = current_mape_df.set_index("KSSS")["current_mape_pct"].to_dict()

# =====================
# Helpers reused from your spec (simplified)
# =====================

def robust_remove_outliers_by_weekday(df: pd.DataFrame, date_col: str, value_col: str = "Чеки") -> pd.DataFrame:
    d = df.copy()
    d = d[[date_col, value_col]].dropna()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce").dt.normalize()
    d[value_col] = pd.to_numeric(d[value_col], errors="coerce")
    d = d.dropna(subset=[date_col, value_col])
    d["weekday"] = d[date_col].dt.weekday
    keep = []
    for _, g in d.groupby("weekday"):
        q1 = g[value_col].quantile(0.25)
        q3 = g[value_col].quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        keep.extend(g[(g[value_col] >= lower) & (g[value_col] <= upper)].index.tolist())
    out = d.loc[sorted(set(keep)), [date_col, value_col]].sort_values(date_col).reset_index(drop=True)
    return out

def compute_weekday_means(window_df: pd.DataFrame, date_col: str, value_col: str = "Чеки") -> Dict[int, float]:
    if window_df.empty:
        return {i: np.nan for i in range(7)}
    wdf = window_df.copy()
    wdf[date_col] = pd.to_datetime(wdf[date_col], errors="coerce").dt.normalize()
    wdf[value_col] = pd.to_numeric(wdf[value_col], errors="coerce")
    wdf = wdf.dropna(subset=[date_col, value_col])
    wdf["weekday"] = wdf[date_col].dt.weekday
    return wdf.groupby("weekday")[value_col].mean().to_dict()

def safe_get_value(daily_df: pd.DataFrame, date: pd.Timestamp, date_col: str, value_col: str) -> Optional[float]:
    row = daily_df.loc[daily_df[date_col] == date]
    if not row.empty:
        return float(row.iloc[0][value_col])
    return None

def mean_ignore_none(values: List[Optional[float]]) -> Optional[float]:
    arr = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
    if not arr:
        return None
    return float(np.mean(arr))

def mape(y_true, y_pred):
    y_true = np.array(y_true, dtype=float)
    y_pred = np.array(y_pred, dtype=float)
    nz = y_true != 0
    if nz.sum() == 0:
        return np.nan
    return float(np.mean(np.abs((y_true[nz] - y_pred[nz]) / y_true[nz])) * 100.0)

def _sarimax_converged(model) -> bool:
    if getattr(model, "converged", None) is False:
        return False
    mle_retvals = getattr(model, "mle_retvals", None)
    if isinstance(mle_retvals, dict) and mle_retvals.get("converged") is False:
        return False
    return True

def build_components_for_month(
    daily_df: pd.DataFrame,
    month_start,
    month_end,
    date_col: str,
    value_col: str,
    cutoff_date: pd.Timestamp,
    use_window_year: bool = True
) -> pd.DataFrame:
    # 1) ограничиваем историю
    hist = daily_df[daily_df[date_col] <= cutoff_date].copy()
    if use_window_year:
        window_start = cutoff_date - pd.DateOffset(years=1) + pd.Timedelta(days=1)
        hist = hist[hist[date_col] >= window_start]
    hist = hist.sort_values(date_col)

    # 2) средние по weekday + общее среднее
    wd_means = compute_weekday_means(hist, date_col=date_col, value_col=value_col)
    overall_mean = float(hist[value_col].mean()) if not hist.empty else np.nan

    # 3) календарь целевого месяца
    days = pd.date_range(month_start, month_end, freq="D")

    rows = []
    for d in days:
        # comp1: тот же день прошлым годом
        same_day_last_year = d - pd.DateOffset(years=1)
        c1 = safe_get_value(hist, same_day_last_year, date_col=date_col, value_col=value_col)

        # comp2: среднее из d-1м и d-2м
        d_prev1 = (d - pd.DateOffset(months=1))
        d_prev2 = (d - pd.DateOffset(months=2))
        c2_list = []
        for cand in (d_prev1, d_prev2):
            if cand <= cutoff_date and (not use_window_year or cand >= (cutoff_date - pd.DateOffset(years=1) + pd.Timedelta(days=1))):
                val = safe_get_value(hist, cand, date_col=date_col, value_col=value_col)
            else:
                val = None
            c2_list.append(val)
        c2 = mean_ignore_none(c2_list)

        # comp3: среднее по weekday
        wd = int(d.weekday())
        c3 = wd_means.get(wd, np.nan)

        # фолбэки
        def fb(x):
            return 0.0 if (x is None or (isinstance(x, float) and math.isnan(x))) else float(x)

        if c1 is None or (isinstance(c1, float) and math.isnan(c1)):
            c1 = c3 if not (isinstance(c3, float) and math.isnan(c3)) else overall_mean
        if c2 is None or (isinstance(c2, float) and math.isnan(c2)):
            c2 = c3 if not (isinstance(c3, float) and math.isnan(c3)) else overall_mean
        if isinstance(c3, float) and math.isnan(c3):
            c3 = overall_mean

        c1, c2, c3 = fb(c1), fb(c2), fb(c3)
        rows.append({"date": d, "comp1": c1, "comp2": c2, "comp3": c3})

    return pd.DataFrame(rows)

def combine_forecast(components_df: pd.DataFrame, w: Tuple[float, float, float]) -> np.ndarray:
    w1, w2, w3 = w
    return (w1 * components_df["comp1"].values +
            w2 * components_df["comp2"].values +
            w3 * components_df["comp3"].values)

# def make_trend_last_dates(ref_end_date: pd.Timestamp, n_days: int = 30) -> List[pd.Timestamp]:
#     """
#     Берём последние n_days до ref_end_date (включая), и переносим их на -1 год.
#     Это аналог periods_maker(...)[1] из старого файла.
#     """
#     trend_now = pd.date_range(end=ref_end_date, periods=n_days+1, freq="D")
#     trend_last = (trend_now - pd.DateOffset(years=1)).normalize()
#     return list(trend_last)

def proportion_of_day_night_df(
    day_night_df: pd.DataFrame,
    ksss: str,
    dates_trend_last: List[pd.Timestamp],
    seg_day: str = "08_20",
    seg_night: str = "other",
    seg_col: str = NAME_OF_COLUMN_OF_SEGMENT,
    value_col: str = VALUE_COL
) -> Tuple[float, float]:
    """
    Возвращает (sum_day, sum_night) по выбранным датам прошлого года.
    """
    d = day_night_df[day_night_df[NAME_OF_COLUMN_OF_KSSS] == ksss].copy()
    if d.empty:
        return 0.0, 0.0

    d = d[d[NAME_OF_COLUMN_OF_DATE].isin(dates_trend_last)]
    if d.empty:
        return 0.0, 0.0

    agg = d.groupby(seg_col)[value_col].sum()
    value_day = float(agg.get(seg_day, 0.0))
    value_night = float(agg.get(seg_night, 0.0))
    return value_day, value_night

# =====================
# ML path helpers
# =====================
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from sklearn.ensemble import RandomForestRegressor
from functions_ml import (
    train_nbeats_forecast,
    train_ttm_global,
    ttm_forecast_series,
    TORCH_AVAILABLE,
    TTM_CONFIG
)

TTM_EVAL_CACHE = None
TTM_FULL_CACHE = None

def ensure_ttm_full_cache():
    global TTM_FULL_CACHE
    if TTM_FULL_CACHE is None:
        TTM_FULL_CACHE = train_ttm_global(
            checks,
            date_col=DATE_COL,
            value_col=VALUE_COL,
            ksss_col=NAME_OF_COLUMN_OF_KSSS,
            config=TTM_CONFIG,
            tail_exclude=0
        )
    return TTM_FULL_CACHE

if TORCH_AVAILABLE:
    try:
        TTM_EVAL_CACHE = train_ttm_global(
            checks,
            date_col=DATE_COL,
            value_col=VALUE_COL,
            ksss_col=NAME_OF_COLUMN_OF_KSSS,
            config=TTM_CONFIG,
            tail_exclude=HORIZON
        )
    except Exception as exc:
        print(f"TTM недоступен: {exc}")

def hampel_filter(series, window_size=15, n_sigmas=3):
    x = series.astype(float).copy()
    med = x.rolling(window=window_size, center=True, min_periods=1).median()
    diff = (x - med).abs()
    mad = diff.rolling(window=window_size, center=True, min_periods=1).median() * 1.4826
    thr = n_sigmas * mad.replace(0, np.nan)
    mask = diff > thr
    x[mask] = med[mask]
    return x

def add_calendar_features(df, date_col="Дата"):
    dt = df[date_col]
    df["day_of_week"] = dt.dt.weekday + 1
    df["is_weekend"] = (df["day_of_week"] >= 6).astype(int)
    df["day_of_month"] = dt.dt.day
    df["month"] = dt.dt.month
    df["week_of_year"] = dt.dt.isocalendar().week.astype(int)
    df["day_of_year"] = dt.dt.dayofyear
    df["dow_occurrence_in_month"] = ((df["day_of_month"] - 1) // 7) + 1
    df["sin_dow"] = np.sin(2 * np.pi * (df["day_of_week"] - 1) / 7)
    df["cos_dow"] = np.cos(2 * np.pi * (df["day_of_week"] - 1) / 7)
    df["sin_month"] = np.sin(2 * np.pi * (df["month"] - 1) / 12)
    df["cos_month"] = np.cos(2 * np.pi * (df["month"] - 1) / 12)
    return df

def make_supervised(df, target_col="Чеки", lags=(1,7,14), roll_windows=(7,14)):
    out = df.copy()
    for L in lags:
        out[f"lag_{L}"] = out[target_col].shift(L)
    for w in roll_windows:
        out[f"roll_mean_{w}"] = out[target_col].shift(1).rolling(w).mean()
        out[f"roll_std_{w}"] = out[target_col].shift(1).rolling(w).std()
    out = out.dropna()
    feat_cols = [c for c in out.columns if c not in [target_col, "Дата"]]
    return out, feat_cols

def evaluate_ml_models(series_df: pd.DataFrame,
                       date_col: str = DATE_COL,
                       value_col: str = VALUE_COL,
                       horizon: int = HORIZON,
                       ksss: Optional[str] = None,
                       ttm_cache=None) -> Tuple[Optional[str], float, Dict[str, float]]:
    df = series_df[[date_col, value_col]].copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.normalize()
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    df = df.dropna(subset=[date_col, value_col]).sort_values(date_col)

    if df.empty or len(df) <= horizon:
        return None, np.nan, {}

    full = pd.date_range(df[date_col].min(), df[date_col].max(), freq="D")
    df = df.set_index(date_col).reindex(full).rename_axis("Дата").reset_index()
    df = df.rename(columns={value_col: "Чеки"})
    if df["Чеки"].isna().any():
        df["Чеки"] = df["Чеки"].interpolate(method="linear").ffill().bfill()
    df["Чеки"] = hampel_filter(df["Чеки"], window_size=15, n_sigmas=3)

    df = add_calendar_features(df, date_col="Дата")
    train_df = df.iloc[:-horizon].copy()
    test_df = df.iloc[-horizon:].copy()
    y_test = test_df["Чеки"].values
    train_series = train_df.set_index("Дата")["Чеки"].astype(float)
    try:
        train_series = train_series.asfreq("D")
    except ValueError:
        pass
    if train_series.isna().any():
        train_series = train_series.ffill().bfill()

    results = {}

    # HoltWinters
    try:
        hw = ExponentialSmoothing(
            train_df["Чеки"], trend="add", seasonal="mul",
            seasonal_periods=SEASONAL_PERIOD, initialization_method="estimated"
        ).fit(optimized=True, use_brute=True)
        y_pred_hw = hw.forecast(horizon).values
        results["HoltWinters"] = mape(y_test, y_pred_hw)
    except Exception:
        results["HoltWinters"] = np.nan

    # SARIMAX
    try:
        best_model, best_score, best_aic, best_pred = None, np.inf, np.inf, None
        p_vals, d_vals, q_vals = [0, 1, 2], [0, 1], [0, 1, 2]
        P_vals, D_vals, Q_vals = [0, 1], [0, 1], [0, 1]
        for p in p_vals:
            for d in d_vals:
                for q in q_vals:
                    for P in P_vals:
                        for D in D_vals:
                            for Q in Q_vals:
                                try:
                                    sar = SARIMAX(
                                        train_series,
                                        order=(p, d, q),
                                        seasonal_order=(P, D, Q, SEASONAL_PERIOD),
                                        enforce_stationarity=False,
                                        enforce_invertibility=False
                                    ).fit(disp=False)
                                    if not _sarimax_converged(sar):
                                        continue
                                    y_pred = sar.forecast(horizon).values
                                    score = mape(y_test, y_pred)
                                    if np.isfinite(score):
                                        if score < best_score:
                                            best_score = score
                                            best_aic, best_model = sar.aic, sar
                                            best_pred = y_pred
                                    elif best_model is None and sar.aic < best_aic:
                                        best_aic, best_model = sar.aic, sar
                                        best_pred = y_pred
                                except Exception:
                                    pass
        if best_model is not None and best_pred is not None:
            y_pred_sar = best_pred
            results["SARIMAX"] = mape(y_test, y_pred_sar)
        else:
            results["SARIMAX"] = np.nan
    except Exception:
        results["SARIMAX"] = np.nan

    # RandomForest
    try:
        base_cols = ["Дата", "Чеки", "day_of_week", "is_weekend", "day_of_month", "month",
                     "week_of_year", "day_of_year", "dow_occurrence_in_month",
                     "sin_dow", "cos_dow", "sin_month", "cos_month"]
        sup, feat_cols = make_supervised(df[base_cols].copy(), target_col="Чеки")
        cutoff_date = test_df["Дата"].min()
        sup_train = sup[sup["Дата"] < cutoff_date].copy()

        rf = RandomForestRegressor(n_estimators=500, random_state=42, n_jobs=-1)
        rf.fit(sup_train[feat_cols], sup_train["Чеки"])

        history = sup[sup["Дата"] < cutoff_date][["Дата", "Чеки"]].copy()
        rf_preds = []
        for i in range(horizon):
            current_date = cutoff_date + pd.Timedelta(days=i)
            row = df[df["Дата"] == current_date][
                ["Дата", "day_of_week", "is_weekend", "day_of_month", "month",
                 "week_of_year", "day_of_year", "dow_occurrence_in_month",
                 "sin_dow", "cos_dow", "sin_month", "cos_month"]
            ].copy()
            if row.empty:
                rf_preds.append(np.nan)
                continue
            hist_series = history.set_index("Дата")["Чеки"]
            for L in (1, 7, 14):
                row[f"lag_{L}"] = hist_series.reindex([current_date - pd.Timedelta(days=L)]).values
            for w in (7, 14):
                roll = hist_series.loc[:current_date - pd.Timedelta(days=1)].tail(w)
                row[f"roll_mean_{w}"] = roll.mean() if len(roll) else np.nan
                row[f"roll_std_{w}"] = roll.std(ddof=0) if len(roll) > 1 else 0.0
            for c in feat_cols:
                if c not in row.columns:
                    row[c] = 0.0
            row = row[feat_cols].fillna(method="ffill").fillna(method="bfill").fillna(0.0)
            y_hat = rf.predict(row)[0]
            rf_preds.append(y_hat)
            history = pd.concat(
                [history, pd.DataFrame({"Дата": [current_date], "Чеки": [y_hat]})],
                ignore_index=True
            )
        results["RandomForest"] = mape(y_test, rf_preds)
    except Exception:
        results["RandomForest"] = np.nan

    # TTM (global)
    if ttm_cache is not None and ksss is not None:
        try:
            y_pred_ttm = ttm_forecast_series(
                ttm_cache,
                train_df["Чеки"].values,
                ksss,
                horizon=horizon
            )
            results["TTM"] = mape(y_test, y_pred_ttm)
        except Exception:
            results["TTM"] = np.nan

    # N-BEATS
    if TORCH_AVAILABLE:
        try:
            nb_res = train_nbeats_forecast(
                series_train=train_df["Чеки"].values,
                horizon=horizon,
                seasonal_period=SEASONAL_PERIOD
            )
            y_pred_nb = nb_res["forecast"]
            results["NBEATS"] = mape(y_test, y_pred_nb)
        except Exception:
            results["NBEATS"] = np.nan
    else:
        results["NBEATS"] = np.nan

    available = {k: v for k, v in results.items() if pd.notna(v)}
    if not available:
        return None, np.nan, results
    best_model = min(available, key=lambda k: available[k])
    return best_model, float(available[best_model]), results

def forecast_ml_for_month(series_df: pd.DataFrame,
                          chosen_model: str,
                          month_start: pd.Timestamp=None,
                          month_end: pd.Timestamp=None,
                          value_col: str=VALUE_COL,
                          period_last_season_override: Optional[List[pd.Timestamp]] = None,
                          ttm_cache=None
                          ) -> pd.DataFrame:
    """
    series_df: columns [date_col, value_col], sorted by date, daily grid (we'll ensure it)
    Returns DataFrame with dates in November and forecast values.
    """
    df = series_df.copy()
    ksss = df[NAME_OF_COLUMN_OF_KSSS].iloc[0]
    # ensure daily grid
    last_obs = pd.to_datetime(series_df[DATE_COL]).max()
    full = pd.date_range(pd.to_datetime(series_df[DATE_COL]).min(), last_obs, freq="D")
    df = df.set_index(DATE_COL).reindex(full).rename_axis(DATE_COL).reset_index()
    # interpolate
    if df[value_col].isna().any():
        df[value_col] = df[value_col].interpolate(method="linear").ffill().bfill()
    # outliers -> Hampel
    df[value_col] = hampel_filter(df[value_col], window_size=15, n_sigmas=3)
    # build features
    base = add_calendar_features(df[[DATE_COL]].copy(), date_col=DATE_COL)
    base[value_col] = df[value_col].values
    last_date = base[DATE_COL].max()

    if month_start == None:
        month_start = periods_maker(df, last_date)[2][0]
        month_end = periods_maker(df, last_date)[2][-1]

    if period_last_season_override is not None:
        period_last_season = period_last_season_override
    else:
        period_last_season = periods_maker(df, last_date)[3]

    # We need predictions for all days in target [month_start..month_end]
    # We'll predict recursively from last_date up to month_end
    start_pred = max(last_date + pd.Timedelta(days=1), month_start)
    horizon = (month_end - start_pred).days + 1
    if horizon <= 0:
        # if we already have data beyond month_end, re-fit and roll forward from (month_start - 1) to month_end
        start_pred = month_start
        horizon = (month_end - start_pred).days + 1

    if chosen_model == "HoltWinters":
        final = ExponentialSmoothing(
            base[value_col], trend="add", seasonal="mul",
            seasonal_periods=SEASONAL_PERIOD, initialization_method="estimated"
        ).fit(optimized=True, use_brute=True)
        future_pred = final.forecast(horizon)
        future_dates = pd.date_range(start_pred, periods=horizon, freq="D")

        df_fc = pd.DataFrame({"Дата": future_dates, "Прогноз": future_pred.values})

    elif chosen_model == "SARIMAX":
        # small grid to find decent params on full data
        best_model, best_aic = None, np.inf
        p_vals, d_vals, q_vals = [0,1,2], [0,1], [0,1,2]
        P_vals, D_vals, Q_vals = [0,1], [0,1], [0,1]
        y = base.set_index(DATE_COL)[value_col].astype(float)
        try:
            y = y.asfreq("D")
        except ValueError:
            pass
        if y.isna().any():
            y = y.ffill().bfill()
        for p in p_vals:
            for d in d_vals:
                for q in q_vals:
                    for P in P_vals:
                        for D in D_vals:
                            for Q in Q_vals:
                                try:
                                    sar = SARIMAX(
                                        y,
                                        order=(p,d,q),
                                        seasonal_order=(P,D,Q,SEASONAL_PERIOD),
                                        enforce_stationarity=False,
                                        enforce_invertibility=False
                                    ).fit(disp=False)
                                    if not _sarimax_converged(sar):
                                        continue
                                    if sar.aic < best_aic:
                                        best_aic, best_model = sar.aic, sar
                                except Exception:
                                    pass
        if best_model is None:
            # fallback to simple HW
            final = ExponentialSmoothing(
                base[value_col], trend="add", seasonal="mul",
                seasonal_periods=SEASONAL_PERIOD, initialization_method="estimated"
            ).fit(optimized=True, use_brute=True)
            future_pred = final.forecast(horizon)
        else:
            future_pred = best_model.forecast(horizon)
        future_dates = pd.date_range(start_pred, periods=horizon, freq="D")
        df_fc = pd.DataFrame({"Дата": future_dates, "Прогноз": np.array(future_pred).astype(float)})

    elif chosen_model == "NBEATS":
        nb_res = train_nbeats_forecast(
            series_train=base[value_col].values,
            horizon=horizon,
            seasonal_period=SEASONAL_PERIOD
        )
        future_pred = nb_res["forecast"]
        future_dates = pd.date_range(start_pred, periods=horizon, freq="D")
        df_fc = pd.DataFrame({"Дата": future_dates, "Прогноз": np.array(future_pred).astype(float)})

    elif chosen_model == "TTM":
        if ttm_cache is None:
            raise ValueError("TTM модель не передана.")
        future_pred = ttm_forecast_series(
            ttm_cache,
            base[value_col].values,
            ksss,
            horizon=horizon
        )
        future_pred = np.asarray(future_pred, dtype=float)
        if len(future_pred) < horizon:
            final = ExponentialSmoothing(
                base[value_col], trend="add", seasonal="mul",
                seasonal_periods=SEASONAL_PERIOD, initialization_method="estimated"
            ).fit(optimized=True, use_brute=True)
            future_pred = np.asarray(final.forecast(horizon), dtype=float)
        elif len(future_pred) > horizon:
            future_pred = future_pred[:horizon]
        future_dates = pd.date_range(start_pred, periods=horizon, freq="D")
        df_fc = pd.DataFrame({"Дата": future_dates, "Прогноз": np.array(future_pred).astype(float)})

    else:  # RandomForest
        base_ml = add_calendar_features(df[[NAME_OF_COLUMN_OF_DATE]].copy(), date_col=NAME_OF_COLUMN_OF_DATE)
        base_ml[value_col] = df[value_col].values
        sup, feat_cols = make_supervised(base_ml.copy(), target_col=value_col)
        if DATE_COL in feat_cols:
            feat_cols.remove(DATE_COL)
        # train on all available
        rf = RandomForestRegressor(n_estimators=500, random_state=42, n_jobs=-1)
        rf.fit(sup[feat_cols], sup[value_col])

        # recursive forecast
        history = base_ml[[NAME_OF_COLUMN_OF_DATE, value_col]].copy()
        preds = []
        dates = []
        cur = start_pred
        while cur <= month_end:
            row = pd.DataFrame({NAME_OF_COLUMN_OF_DATE:[cur]})
            row = add_calendar_features(row, date_col=NAME_OF_COLUMN_OF_DATE)

            # build lags from history
            hist = history.set_index(NAME_OF_COLUMN_OF_DATE)[value_col]
            for L in (1,7,14):
                row[f"lag_{L}"] = hist.reindex([cur - pd.Timedelta(days=L)]).values
            for w in (7,14):
                roll = hist.loc[:cur - pd.Timedelta(days=1)].tail(w)
                row[f"roll_mean_{w}"] = roll.mean() if len(roll) else np.nan
                row[f"roll_std_{w}"]  = roll.std(ddof=0) if len(roll) > 1 else 0.0
            for c in feat_cols:
                if c not in row.columns:
                    row[c] = 0.0
            row = row[feat_cols].fillna(method="ffill").fillna(method="bfill").fillna(0.0)
            y_hat = rf.predict(row)[0]
            preds.append(y_hat)
            dates.append(cur)
            # update history
            history = pd.concat([history, pd.DataFrame({NAME_OF_COLUMN_OF_DATE:[cur], value_col:[y_hat]})], ignore_index=True)
            cur += pd.Timedelta(days=1)
        df_fc = pd.DataFrame({"Дата": dates, "Прогноз": np.array(preds).astype(float)})
    # keep only target month
    df_fc[NAME_OF_COLUMN_OF_KSSS] = ksss
    df_fc = df_fc[(df_fc["Дата"] >= month_start) & (df_fc["Дата"] <= month_end)].copy()
    val_day, val_night = proportion_of_day_night_df(day_night, ksss, period_last_season)
    # если нет нормальных дневных/ночных долей — сигналим наверх
    if val_day + val_night == 0:
        last_date -= timedelta(days=int(year_len(now.year)))
        period_last_season = periods_maker(df, last_date)[3]
        val_day, val_night = proportion_of_day_night_df(day_night, ksss, period_last_season)
        if val_day + val_night == 0:
            last_date = base[DATE_COL].max()
            period_last_season = periods_maker(df, last_date)[0]
            val_day, val_night = proportion_of_day_night_df(day_night, ksss, period_last_season)
            if val_day + val_night == 0:
                val_day, val_night = 0.77, 0.23

    df_fc['Чеки дневные'] = df_fc['Прогноз']*(val_day/(val_day+val_night))
    df_fc['Чеки ночные'] = df_fc['Прогноз']*(val_night/(val_day+val_night))
    df_fc = df_fc.rename(columns={'Дата': 'date'})
    df_fc[["Прогноз", "Чеки дневные", "Чеки ночные"]] = df_fc[["Прогноз", "Чеки дневные", "Чеки ночные"]].round(0)
    df_fc = df_fc[["ORG_KSSS", "date", "Прогноз", "Чеки дневные", "Чеки ночные"]]
    return df_fc


def current_forecast_for_ksss(df: pd.DataFrame, periods=None) -> pd.DataFrame:
    """
    Прогноз по current-логике для одного KSSS.
    df должен содержать колонки: 'date', 'unique_cheques', 'ORG_KSSS'
    Возвращает df с колонками:
    ['ORG_KSSS','date','Чеки сутки (шт.)','Чеки день (шт.)','Чеки ночь (шт.)']
    """
    if periods is None:
        dates_trend_now, dates_trend_last, dates_season_now, dates_season_last = periods_maker(df)
    else:
        dates_trend_now, dates_trend_last, dates_season_now, dates_season_last = periods

    df_t1 = df[df['date'].isin(dates_trend_last)]
    df_t0 = df[df['date'].isin(dates_trend_now)]
    df_s1 = df[df['date'].isin(dates_season_last)]

    def bad_count(df_):
        s = df_['unique_cheques']
        return s.isna().sum() + (s == 0).sum()

    if bad_count(df_t0) > 1 or bad_count(df_t1) > 1 or bad_count(df_s1) > 1:
        return pd.DataFrame()


    sum_t0 = df_t0['unique_cheques'].sum()
    sum_t1 = df_t1['unique_cheques'].sum()
    sum_s1 = df_s1['unique_cheques'].sum()

    if any((x == 0) or pd.isna(x) for x in (sum_t0, sum_t1, sum_s1)):
        return pd.DataFrame()

    sum_s0 = round(sum_s1 * sum_t1 / sum_t0, 0)

    # --- дальше твоя логика 1-в-1 ---
    df_ksss_forc_prev = df_s1.copy()
    df_ksss_forc_prev = indicator_adder(df_ksss_forc_prev, sum_s1)

    idx = pd.date_range(start=dates_season_now[0], end=dates_season_now[-1], freq="D")
    df_ksss_forc = pd.DataFrame(index=idx)
    ksss = df_ksss_forc_prev[NAME_OF_COLUMN_OF_KSSS].iloc[0]
    df_ksss_forc['ORG_KSSS'] = ksss
    df_ksss_forc["year"] = df_ksss_forc.index.year
    df_ksss_forc["month"] = df_ksss_forc.index.month
    df_ksss_forc["day"] = df_ksss_forc.index.day
    df_ksss_forc["date"] = df_ksss_forc.index
    df_ksss_forc = indicator_adder(df_ksss_forc, sum_s0)

    df_ksss_forc = df_ksss_forc.merge(
        df_ksss_forc_prev[['Сцепка', 'Доля, %']],
        on='Сцепка',
        how='left'
    )

    df_weekday_avg = (
        df_ksss_forc_prev
        .groupby('ДенНед', sort=False)['Доля, %']
        .mean()
        .rename('Доля, %_weekday_mean')
    )

    df_ksss_forc = df_ksss_forc.merge(
        df_weekday_avg,
        left_on='ДенНед',
        right_index=True,
        how='left'
    )

    df_ksss_forc['Доля, %'] = df_ksss_forc['Доля, %'].fillna(df_ksss_forc['Доля, %_weekday_mean'])
    df_ksss_forc['Доля, %'] = df_ksss_forc['Доля, %'].fillna(df_ksss_forc_prev['Доля, %'].mean())
    df_ksss_forc.drop(columns=['Доля, %_weekday_mean'], inplace=True)

    df_ksss_forc['Прогноз'] = round(df_ksss_forc['Доля, %'] * sum_s0, 0)

    value_day, value_night = proportion_of_day_night(PATH_TO_DAY_NIGHT_FILE, ksss, dates_trend_last)
    if value_day + value_night == 0:
        return pd.DataFrame()

    df_ksss_forc['Чеки дневные'] = round(df_ksss_forc['Прогноз'] * (value_day/(value_day+value_night)), 0)
    df_ksss_forc['Чеки ночные'] = round(df_ksss_forc['Прогноз'] * (value_night/(value_day+value_night)), 0)

    df_ksss_forc = df_ksss_forc[['date', 'ORG_KSSS', 'Прогноз', 'Чеки дневные', 'Чеки ночные']]
    df_ksss_forc = df_ksss_forc.rename(columns={'Дата': 'date'})
    df_ksss_forc = df_ksss_forc[["ORG_KSSS", "date", "Прогноз", "Чеки дневные", "Чеки ночные"]]
    return df_ksss_forc.reset_index(drop=True)

def build_periods_for_range(df: pd.DataFrame,
                            start_date: pd.Timestamp,
                            end_date: pd.Timestamp) -> Tuple[List[pd.Timestamp], List[pd.Timestamp], List[pd.Timestamp], List[pd.Timestamp]]:
    dates_trend_now, dates_trend_last, _, _ = periods_maker(df)
    dates_season_now = pd.date_range(start_date, end_date, freq="D").to_list()
    dates_season_last = [(d - pd.DateOffset(years=1)).normalize() for d in dates_season_now]
    return dates_trend_now, dates_trend_last, dates_season_now, dates_season_last

def forecast_weights_for_ksss(df_k: pd.DataFrame,
                              ksss: str,
                              w1: float,
                              w2: float,
                              w3: float,
                              month_start: pd.Timestamp,
                              month_end: pd.Timestamp,
                              period_of_month: List[pd.Timestamp],
                              one_year_ago: timedelta) -> pd.DataFrame:
    clean_df = robust_remove_outliers_by_weekday(df_k, date_col=NAME_OF_COLUMN_OF_DATE, value_col=VALUE_COL)
    cutoff = month_start - pd.Timedelta(days=45)
    comps = build_components_for_month(
        daily_df=clean_df,
        month_start=month_start,
        month_end=month_end,
        date_col=NAME_OF_COLUMN_OF_DATE,
        value_col=VALUE_COL,
        cutoff_date=cutoff,
        use_window_year=True
    )
    comps = comps[comps["date"].between(month_start, month_end)]
    if comps.empty:
        return pd.DataFrame()

    pred = combine_forecast(comps, (w1, w2, w3))
    fc_df = pd.DataFrame({"Дата": comps["date"], "KSSS": ksss, "Прогноз": np.round(pred, 0).astype(int)})

    value_day, value_night = proportion_of_day_night_df(day_night, ksss, period_of_month)
    if value_day + value_night == 0:
        last_year_period = make_timestamp_range(
            month_start - one_year_ago,
            month_end - one_year_ago,
            freq="D",
            inclusive="both"
        )
        value_day, value_night = proportion_of_day_night_df(day_night, ksss, last_year_period)
        if value_day + value_night == 0:
            value_day, value_night = 0.77, 0.23

    fc_df['Чеки дневные'] = round(fc_df['Прогноз'] * (value_day / (value_day + value_night)), 0)
    fc_df['Чеки ночные'] = round(fc_df['Прогноз'] * (value_night / (value_day + value_night)), 0)
    fc_df = fc_df.rename(columns={"Дата": "date", "KSSS": "ORG_KSSS"})
    return fc_df[["ORG_KSSS", "date", "Прогноз", "Чеки дневные", "Чеки ночные"]]

# =====================
# Main loop per KSSS
# =====================
last_date_all = checks['date'].iloc[-1]

def first_last_dates(last_date_all):
    y = (last_date_all + timedelta(days=HORIZON - 15)).year
    m = (last_date_all + timedelta(days=HORIZON - 15)).month
    MONTH_LEN = calendar.monthrange(y, m)[-1]
    first = last_date_all + timedelta(days=HORIZON+DAYS_UNTIL_SENDINGS-int(MONTH_LEN))
    last = last_date_all + timedelta(days=HORIZON + DAYS_UNTIL_SENDINGS+(30-MONTH_LEN))
    return first, last

def make_timestamp_range(start, end, freq="D", inclusive="both"):
    start = pd.to_datetime(start)
    end   = pd.to_datetime(end)
    return pd.date_range(start, end, freq=freq, inclusive=inclusive).to_list()

def first_last_dates_calendar_month(last_date_all, horizon_days, sendings_days):
    """
    last_date_all: последняя дата факта
    horizon_days: HORIZON
    sendings_days: DAYS_UNTIL_SENDINGS

    Возвращает границы целевого календарного месяца:
    [первый день месяца, последний день месяца]
    """
    # дата, относительно которой выбираем месяц прогноза
    ref = pd.to_datetime(last_date_all) + timedelta(days=horizon_days + sendings_days)

    # первый день целевого месяца
    first = ref.replace(day=1)

    # последний день целевого месяца
    last_day = calendar.monthrange(first.year, first.month)[1]
    last = first.replace(day=last_day)

    return first.normalize(), last.normalize()

def append_to_csv(df: pd.DataFrame, path: str) -> None:
    if df is None or df.empty:
        return
    header = not os.path.exists(path)
    df.to_csv(path, mode="a", index=False, header=header, encoding="utf-8")


def _format_number(value):
    if pd.isna(value):
        return None
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)):
        if float(value).is_integer():
            return int(value)
        return round(float(value), 1)
    return value


def build_recommendations_export(
    forecast_path: str | Path,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
    now: dt.datetime,
) -> Optional[Path]:
    template_path = (
        Path(__file__).resolve().parents[1]
        / "Проверка чеков"
        / "Анализ рассылок"
        / "Рассылки по месяцам - все"
        / "01.26 Январь 2026.xlsx"
    )
    if not template_path.exists():
        print(f"Не найден шаблон рассылки: {template_path}")
        return None
    forecast_path = Path(forecast_path)
    if not forecast_path.exists():
        print(f"Не найден файл прогноза для рассылки: {forecast_path}")
        return None
    try:
        forecast_df = pd.read_csv(forecast_path)
    except Exception as exc:
        print(f"Не удалось прочитать прогноз для рассылки: {exc}")
        return None
    if forecast_df.empty:
        print("Прогноз пустой, рассылка не сформирована.")
        return None

    required_cols = {
        "ORG_KSSS",
        "date",
        "Чеки дневные",
        "Чеки ночные",
        "Часы дневные",
        "Часы ночные",
        "Совокупное кол-во суточных чеков",
        "Совокупное кол-во суточных часов",
    }
    missing = [col for col in required_cols if col not in forecast_df.columns]
    if missing:
        print(f"Не хватает колонок для рассылки: {missing}")
        return None

    forecast_df = forecast_df.copy()
    forecast_df["ORG_KSSS"] = forecast_df["ORG_KSSS"].apply(normalize_ksss)
    forecast_df["date"] = pd.to_datetime(forecast_df["date"], errors="coerce")
    forecast_df = forecast_df.dropna(subset=["date"])
    forecast_df = forecast_df[(forecast_df["date"] >= month_start) & (forecast_df["date"] <= month_end)]
    if forecast_df.empty:
        print("Нет данных прогноза в целевом месяце для рассылки.")
        return None
    forecast_df["day"] = forecast_df["date"].dt.day

    cls_path = Path(__file__).resolve().parents[1] / "cls_2025_12_AZS.xlsx"
    cls_cols = [
        "КССС_union",
        "НПО",
        "ФИО_РУ",
        "ФИО_ТМ_Агента",
        "ФИО_менеджер_АЗС",
        "Название_АЗС",
    ]
    meta_df = pd.DataFrame(columns=["KSSS_norm", "НПО", "РУ", "ТМ", "Менеджер АЗС", "Номер АЗС"])
    if cls_path.exists():
        try:
            cls_df = pd.read_excel(cls_path, sheet_name="cls_AZS", usecols=cls_cols)
        except ValueError:
            cls_df = pd.read_excel(cls_path, sheet_name="cls_AZS")
        missing_meta = [col for col in cls_cols if col not in cls_df.columns]
        if missing_meta:
            print(f"Не хватает колонок в cls_2025_12_AZS.xlsx: {missing_meta}")
        else:
            cls_df = cls_df[cls_cols].rename(columns={
                "КССС_union": "KSSS_norm",
                "ФИО_РУ": "РУ",
                "ФИО_ТМ_Агента": "ТМ",
                "ФИО_менеджер_АЗС": "Менеджер АЗС",
                "Название_АЗС": "Номер АЗС",
            })
            cls_df["KSSS_norm"] = cls_df["KSSS_norm"].apply(normalize_ksss)
            meta_df = cls_df.dropna(subset=["KSSS_norm"]).drop_duplicates("KSSS_norm")
    else:
        print(f"Не найден файл справочника АЗС: {cls_path}")

    forecast_df = forecast_df.merge(meta_df, left_on="ORG_KSSS", right_on="KSSS_norm", how="left")

    month_name = MONTH_NAMES.get(month_start.month, str(month_start.month))
    header_title = (
        "Рекомендации к планированию графиков сменности для сотрудников "
        f"в должности 'Продавец', 'Старший продавец', 'Оператор АЗС' на {month_name} {month_start.year}"
    )
    wb = load_workbook(template_path)
    ws = wb.active

    days_in_month = month_len(month_start.year, month_start.month)
    day_columns = []
    for cell in ws[3]:
        if isinstance(cell.value, int):
            day_columns.append((int(cell.value), cell.column))
    cols_to_delete = sorted([col for day, col in day_columns if day > days_in_month], reverse=True)
    for col in cols_to_delete:
        ws.delete_cols(col, 1)

    ws["A1"].value = header_title

    day_col_map = {}
    for cell in ws[3]:
        if isinstance(cell.value, int):
            day_col_map[int(cell.value)] = cell.column

    for day, col in day_col_map.items():
        ws.cell(3, col).value = day
        weekday = None
        if day <= days_in_month:
            weekday = WEEKDAY[dt.date(month_start.year, month_start.month, day).isoweekday()]
        ws.cell(4, col).value = weekday

    header_cells = {cell.value: cell.column for cell in ws[3] if cell.value is not None}
    col_map = {
        "НПО": header_cells.get("НПО"),
        "РУ": header_cells.get("РУ"),
        "ТМ": header_cells.get("ТМ"),
        "Менеджер АЗС": header_cells.get("Менеджер АЗС"),
        "Номер АЗС": header_cells.get("Номер АЗС"),
        "КССС": header_cells.get("КССС"),
        "Показатель": header_cells.get("Показатель"),
        "Всего": header_cells.get("Всего"),
        "Рекомендуемая численность": header_cells.get("Рекомендуемая численность"),
    }
    for key in ("Всего", "Рекомендуемая численность"):
        col_idx = col_map.get(key)
        if col_idx is None:
            continue
        in_merge = False
        for merge_range in ws.merged_cells.ranges:
            if merge_range.min_row <= 3 <= merge_range.max_row and merge_range.min_col <= col_idx <= merge_range.max_col:
                in_merge = True
                break
        if not in_merge:
            ws.merge_cells(start_row=3, end_row=4, start_column=col_idx, end_column=col_idx)

    data_start_row = 5
    max_col = ws.max_column
    style_rows = []
    row_heights = []
    for idx in range(6):
        src_row = data_start_row + idx
        row_styles = []
        for col in range(1, max_col + 1):
            cell = ws.cell(src_row, col)
            row_styles.append({
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            })
        style_rows.append(row_styles)
        row_heights.append(ws.row_dimensions[src_row].height)

    row_merges = [
        r for r in ws.merged_cells.ranges
        if r.min_row == data_start_row and r.max_row == data_start_row
    ]
    message_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and "Просим заполнить анкету" in cell.value:
                message_row = cell.row
                break
        if message_row:
            break
    if message_row is None:
        message_row = ws.max_row + 1

    if message_row > data_start_row:
        ws.delete_rows(data_start_row, message_row - data_start_row)
        message_row = data_start_row

    indicator_map = [
        ("Прогноз дневных чеков (дневная смена)", "Чеки дневные"),
        ("Прогноз ночных чеков (ночная смена)", "Чеки ночные"),
        ("Совокупная сумма часов для дневной смены", "Часы дневные"),
        ("Совокупная сумма часов для ночной смены", "Часы ночные"),
        ("Итого прогноз чеков на сутки", "Совокупное кол-во суточных чеков"),
        ("Итого целевая сумма часов на сутки", "Совокупное кол-во суточных часов"),
    ]

    groups = list(forecast_df.groupby("ORG_KSSS", sort=True))
    total_rows = len(groups) * len(indicator_map)
    blank_rows = 1
    if total_rows > 0:
        ws.insert_rows(message_row, amount=total_rows + blank_rows)

    data_end_row = data_start_row + total_rows - 1
    if total_rows > 0:
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.max_row < data_start_row or merge_range.min_row > data_end_row:
                continue
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row, col)
            ws.unmerge_cells(str(merge_range))

    for g_idx, (ksss, group) in enumerate(groups):
        block_start = data_start_row + g_idx * len(indicator_map)
        block_end = block_start + len(indicator_map) - 1
        meta_row = group[["НПО", "РУ", "ТМ", "Менеджер АЗС", "Номер АЗС"]].iloc[0]
        ksss_value = coerce_int(ksss)
        if ksss_value is None:
            try:
                ksss_value = int(float(str(ksss).strip()))
            except Exception:
                ksss_value = ksss
        for idx, (label, col_name) in enumerate(indicator_map):
            row_idx = block_start + idx
            for col in range(1, max_col + 1):
                style = style_rows[idx][col - 1]
                cell = ws.cell(row_idx, col)
                cell.font = copy(style["font"])
                cell.fill = copy(style["fill"])
                cell.border = copy(style["border"])
                cell.alignment = copy(style["alignment"])
                cell.number_format = style["number_format"]
                cell.protection = copy(style["protection"])
            if row_heights[idx] is not None:
                ws.row_dimensions[row_idx].height = row_heights[idx]

            if col_map["Показатель"]:
                ws.cell(row_idx, col_map["Показатель"]).value = label

            day_series = pd.to_numeric(group.groupby("day", sort=True)[col_name].sum(), errors="coerce")
            for day, col in day_col_map.items():
                if day > days_in_month:
                    ws.cell(row_idx, col).value = None
                    continue
                value = _format_number(day_series.get(day))
                ws.cell(row_idx, col).value = value

            if col_map["Всего"]:
                total_value = _format_number(day_series.sum())
                ws.cell(row_idx, col_map["Всего"]).value = total_value
            if col_map["Рекомендуемая численность"]:
                if idx == len(indicator_map) - 1:
                    total_hours = day_series.sum()
                    if pd.isna(total_hours):
                        recommended = None
                    else:
                        recommended = int(math.floor(float(total_hours) / 164 + 0.5))
                    ws.cell(row_idx, col_map["Рекомендуемая численность"]).value = recommended
                else:
                    ws.cell(row_idx, col_map["Рекомендуемая численность"]).value = None

        for row_idx in range(block_start, block_end + 1):
            if col_map["НПО"]:
                ws.cell(row_idx, col_map["НПО"]).value = meta_row.get("НПО")
            if col_map["РУ"]:
                ws.cell(row_idx, col_map["РУ"]).value = meta_row.get("РУ")
            if col_map["ТМ"]:
                ws.cell(row_idx, col_map["ТМ"]).value = meta_row.get("ТМ")
            if col_map["Менеджер АЗС"]:
                ws.cell(row_idx, col_map["Менеджер АЗС"]).value = meta_row.get("Менеджер АЗС")
            if col_map["Номер АЗС"]:
                ws.cell(row_idx, col_map["Номер АЗС"]).value = meta_row.get("Номер АЗС")
            if col_map["КССС"]:
                ws.cell(row_idx, col_map["КССС"]).value = ksss_value

    if total_rows > 0:
        for row_idx in range(data_start_row, data_end_row + 1):
            for merge_range in row_merges:
                ws.merge_cells(
                    start_row=row_idx,
                    end_row=row_idx,
                    start_column=merge_range.min_col,
                    end_column=merge_range.max_col,
                )

        def _merge_identical(col_start: int, col_end: int) -> None:
            if col_start is None or col_end is None:
                return
            current_value = None
            current_start = data_start_row
            for row_idx in range(data_start_row, data_end_row + 2):
                if row_idx <= data_end_row:
                    value = ws.cell(row_idx, col_start).value
                    if isinstance(value, str) and not value.strip():
                        value = None
                else:
                    value = None
                if value != current_value:
                    if current_value is not None and row_idx - current_start > 1:
                        ws.merge_cells(
                            start_row=current_start,
                            end_row=row_idx - 1,
                            start_column=col_start,
                            end_column=col_end,
                        )
                    current_value = value
                    current_start = row_idx

        ru_col = col_map["РУ"]
        ru_end_col = None
        if ru_col is not None:
            if ws.cell(3, ru_col + 1).value is None:
                ru_end_col = ru_col + 1
            else:
                ru_end_col = ru_col

        merge_targets = [
            (col_map["НПО"], col_map["НПО"]),
            (ru_col, ru_end_col),
            (col_map["ТМ"], col_map["ТМ"]),
            (col_map["Менеджер АЗС"], col_map["Менеджер АЗС"]),
            (col_map["Номер АЗС"], col_map["Номер АЗС"]),
            (col_map["КССС"], col_map["КССС"]),
        ]
        for col_start, col_end in merge_targets:
            _merge_identical(col_start, col_end)

    new_message_row = data_start_row + total_rows + blank_rows

    if ws._images:
        img = ws._images[0]
        try:
            px_side = 120
            img.width = px_side
            img.height = px_side
            new_row = new_message_row - 1
            new_col = 1
            marker = AnchorMarker(col=new_col, colOff=0, row=new_row, rowOff=0)
            emu = px_side * 9525
            img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(cx=emu, cy=emu))
        except Exception:
            pass

    message_text = (
        "Просим заполнить анкету обратной связи о предоставляемых рекомендациях к графикам сменности линейного персонала АЗС.\n"
        "Для заполнения анкеты необходимо отсканировать QR-код с мобильного телефона."
    )
    ws.merge_cells(start_row=new_message_row, end_row=new_message_row, start_column=5, end_column=11)
    message_set = False
    for row in ws.iter_rows(min_row=new_message_row, max_row=min(new_message_row + 1, ws.max_row)):
        for cell in row:
            if isinstance(cell.value, str) and "Просим заполнить анкету" in cell.value:
                cell.value = message_text
                message_set = True
                break
        if message_set:
            break
    if not message_set:
        target_cell = ws.cell(new_message_row, 5)
        target_cell.value = message_text
    target_cell = ws.cell(new_message_row, 5)
    target_cell.font = copy(target_cell.font)
    target_cell.font = target_cell.font.copy(italic=True)
    target_cell.alignment = copy(target_cell.alignment)
    target_cell.alignment = target_cell.alignment.copy(horizontal="left", vertical="bottom", wrap_text=True)
    ws.row_dimensions[new_message_row].height = 35 * 0.75

    if total_rows > 0:
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row < data_end_row + 1:
                continue
            if (
                merge_range.min_row == new_message_row
                and merge_range.max_row == new_message_row
                and merge_range.min_col == 5
                and merge_range.max_col == 11
            ):
                continue
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row, col)
            ws.unmerge_cells(str(merge_range))

    last_keep_row = new_message_row + 1
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)

    export_root = Path(RESULT_PREDICTIONS).parents[1] / "Рекомендации формат"
    export_dir = export_root / f"Рекомендации {month_name} {month_start.year}"
    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = now.strftime("%d.%m.%Y %H.%M")
    export_name = f"{month_start.month:02d}.{month_start.year % 100:02d} {month_name} {month_start.year} ({stamp}).xlsx"
    export_path = export_dir / export_name
    wb.save(export_path)
    return export_path


def print_mape_stats(path: str) -> None:
    try:
        df = pd.read_csv(path)
    except Exception as exc:
        print(f"Не удалось прочитать MAPE файл для статистики: {exc}")
        return
    if df.empty:
        return

    total = len(df)
    print("\n--- Статистика по MAPE ---")
    print(f"Всего строк: {total}")

    if "ML_model" in df.columns:
        ml_series = df["ML_model"].astype(str).str.strip()
        ml_nonempty = ml_series != ""
        ml_total = int(ml_nonempty.sum())
        ttm_ml = int((ml_series == "TTM").sum())
        if ml_total > 0:
            print(f"TTM как лучшая ML-модель: {ttm_ml}/{ml_total} ({round(ttm_ml / ml_total * 100, 1)}%)")
        else:
            print("TTM как лучшая ML-модель: 0/0 (0%)")

    if "Способ прогноза" in df.columns:
        src_series = df["Способ прогноза"].astype(str).str.strip()
        src_counts = src_series.replace("", "NA").value_counts()
        src_info = ", ".join([f"{k}: {v}" for k, v in src_counts.items()])
        print(f"Источники прогноза: {src_info}")

    if "Способ прогноза" in df.columns and "Детали прогноза" in df.columns:
        src_series = df["Способ прогноза"].astype(str).str.strip()
        detail_series = df["Детали прогноза"].astype(str).str.strip()
        ml_used = int((src_series == "МЛ").sum())
        ttm_used = int(((src_series == "МЛ") & (detail_series == "TTM")).sum())
        if ml_used > 0:
            print(f"TTM как фактически выбранный прогноз: {ttm_used}/{ml_used} ({round(ttm_used / ml_used * 100, 1)}%)")
        else:
            print("TTM как фактически выбранный прогноз: 0/0 (0%)")

def finalize_forecast_output(df: pd.DataFrame, hours_calc: "HoursCalculator") -> pd.DataFrame:
    df = df.copy()
    df = hours_calc.add_hours(df, output_day_col="Часы дневные", output_night_col="Часы ночные")
    df["Совокупное кол-во суточных чеков"] = (
        pd.to_numeric(df["Чеки дневные"], errors="coerce").fillna(0)
        + pd.to_numeric(df["Чеки ночные"], errors="coerce").fillna(0)
    ).round(0).astype(int)
    df["Совокупное кол-во суточных часов"] = (
        pd.to_numeric(df["Часы дневные"], errors="coerce").fillna(0)
        + pd.to_numeric(df["Часы ночные"], errors="coerce").fillna(0)
    ).round(0).astype(int)
    return df[[
        "ORG_KSSS",
        "date",
        "Чеки дневные",
        "Чеки ночные",
        "Часы дневные",
        "Часы ночные",
        "Совокупное кол-во суточных чеков",
        "Совокупное кол-во суточных часов"
    ]]

env_start = parse_env_date(FORECAST_START)
env_end = parse_env_date(FORECAST_END)
manual_period = False
if env_start is not None and env_end is not None:
    target_month_start = env_start.normalize()
    target_month_end = env_end.normalize()
    if target_month_end < target_month_start:
        raise ValueError("Дата окончания периода меньше даты начала.")
    manual_period = True
else:
    target_month_start, target_month_end = first_last_dates_calendar_month(
        last_date_all, HORIZON, 0
    )
print(last_date_all)
days_in_year = year_len(target_month_start.year)

one_year_ago = timedelta(days=days_in_year)
period_of_month = make_timestamp_range(target_month_start-one_year_ago, target_month_end-one_year_ago, freq="D", inclusive="both")
ksss_list = sorted(checks[NAME_OF_COLUMN_OF_KSSS].unique().tolist())
skipped_info = []
total_ksss = len(ksss_list)
write_progress(0, total_ksss)

now = dt.datetime.now()
display_month = target_month_start
mnth_int = int(display_month.month)
mnth = MONTH_NAMES.get(mnth_int)
len_mnth = len(mnth)

os.makedirs(RESULT_PREDICTIONS, exist_ok=True)
os.makedirs(OUTPUT_MODELS_COMPARISON, exist_ok=True)
write_result_dir(RESULT_PREDICTIONS)
forecast_file_name = f"Прогноз чеков от {now.day:02d}.{now.month:02d}.{now.year} {now.hour:02d}.{now.minute:02d}.csv"
out_csv = os.path.join(RESULT_PREDICTIONS, forecast_file_name)
mape_file_name = f"MAPE итоговые {now.day:02d}.{now.month:02d}.{now.year} {now.hour:02d}.{now.minute:02d}.csv"
mape_out_csv = os.path.join(OUTPUT_MODELS_COMPARISON, mape_file_name)

if HoursCalculator is None:
    detail = ""
    if _hours_import_path is not None:
        detail = f" Путь: {_hours_import_path}."
    if _hours_import_error is not None:
        detail = f"{detail} Ошибка: {_hours_import_error}."
    raise ImportError("Не удалось импортировать Hours_from_checks_simple.HoursCalculator." + detail)
hours_calculator = HoursCalculator()

print(52*'-')
print(f'------------------- \033[1m Добрый день! \033[0m -----------------')
print(f'--------------- \033[1m Сегодня {now.day:02d}.{now.month:02d}.{now.year} \033[0m ---------------')
print(f'-------- Приступаем к формированию прогнозов -------')
print(f'----------- на \033[1m{mnth.upper()}\033[0m для \033[1m{len(ksss_list)}\033[0m объектов -----------')
print(52*'-')


start_all = time.time()
in_work += (start_all - beginning_time)
forecasted_ksss = 0
for index, k in enumerate(ksss_list):
    start = time.time()
    if stop_requested():
        print("Получен сигнал остановки. Завершаем работу.")
        break
    try:
        print(f'\n\033[4m{index+1}/{len(sp_final)}\033[0m ({round(int(index+1)/len(sp_final)*100, 1)}%)\n')
        print(f'Работаем над прогнозом для АЗС с КССС: \033[1m{k}\033[0m')
        df_k = checks.loc[
            checks[NAME_OF_COLUMN_OF_KSSS] == k,
            [NAME_OF_COLUMN_OF_KSSS, DATE_COL, VALUE_COL]
        ].copy()

        df_k = df_k.rename(columns={
            DATE_COL: "date",
            VALUE_COL: "unique_cheques"
        }).dropna()

        df_k = df_k.sort_values("date").reset_index(drop=True)

        if df_k.empty:
            skipped_info.append((k, "Нет данных чеков"))
            continue

        k_key = normalize_ksss(k)
        weight_info = weights_map.get(k_key)
        w1 = w2 = w3 = np.nan
        mape_weights = np.nan
        if weight_info:
            w1 = weight_info.get("w1", np.nan)
            w2 = weight_info.get("w2", np.nan)
            w3 = weight_info.get("w3", np.nan)
            mape_weights = weight_info.get("weights_mape_pct", np.nan)

        mape_cur = current_mape_map.get(k_key, np.nan)
        best_ml_model, mape_ml, _ = evaluate_ml_models(
            df_k,
            date_col=DATE_COL,
            value_col=VALUE_COL,
            horizon=HORIZON,
            ksss=k,
            ttm_cache=TTM_EVAL_CACHE
        )
        periods_override = build_periods_for_range(df_k, target_month_start, target_month_end) if manual_period else None

        mape_candidates = {
            "weights": mape_weights,
            "ml": mape_ml,
            "current": mape_cur,
        }
        available = {src: val for src, val in mape_candidates.items() if pd.notna(val)}
        if not available:
            mape_row = pd.DataFrame([{
                "KSSS": k,
                "MAPE_ML": round(mape_ml, 3) if pd.notna(mape_ml) else np.nan,
                "ML_model": best_ml_model or "",
                "MAPE_weights": round(mape_weights, 3) if pd.notna(mape_weights) else np.nan,
                "MAPE_cur": round(mape_cur, 3) if pd.notna(mape_cur) else np.nan,
                "Способ прогноза": "",
                "Детали прогноза": ""
            }])
            append_to_csv(mape_row, mape_out_csv)
            skipped_info.append((k, "Нет MAPE для выбора модели"))
            continue

        source_priority = {"weights": 0, "ml": 1, "current": 2}
        ranked_sources = sorted(available.items(), key=lambda kv: (kv[1], source_priority[kv[0]]))

        fc_df = pd.DataFrame()
        chosen_source = None
        chosen_model = None
        chosen_mape = np.nan

        for src, mape_val in ranked_sources:
            if src == "weights":
                if weight_info is None or pd.isna(w1) or pd.isna(w2) or pd.isna(w3):
                    continue
                fc_df = forecast_weights_for_ksss(
                    df_k=df_k,
                    ksss=k,
                    w1=float(w1),
                    w2=float(w2),
                    w3=float(w3),
                    month_start=target_month_start,
                    month_end=target_month_end,
                    period_of_month=period_of_month,
                    one_year_ago=one_year_ago
                )
                chosen_model = "weights"
            elif src == "ml":
                if not best_ml_model:
                    continue
                ttm_cache = None
                if best_ml_model == "TTM":
                    ttm_cache = ensure_ttm_full_cache()
                fc_df = forecast_ml_for_month(
                    df_k,
                    best_ml_model,
                    target_month_start,
                    target_month_end,
                    value_col=VALUE_COL,
                    period_last_season_override=period_of_month,
                    ttm_cache=ttm_cache
                )
                chosen_model = best_ml_model
            else:
                fc_df = current_forecast_for_ksss(df_k, periods_override)
                chosen_model = "current"

            if fc_df is not None and not fc_df.empty:
                chosen_source = src
                chosen_mape = mape_val
                break

        if fc_df is None or fc_df.empty:
            skipped_info.append((k, "Не удалось получить прогноз ни по одному источнику"))
            continue

        print(f'Выбраная модель прогноза: {chosen_model}')
        print(f"Прогнозирование происходит методом: \033[1m{chosen_source}\033[0m")
        if chosen_source == "weights":
            print(f'Вес сезонного месяца: \033[1m{w1}\033[0m')
            print(f'Вес тренда последних двух месяцев: \033[1m{w2}\033[0m')
            print(f'Вес среднего по дню недели: \033[1m{w3}\033[0m')
        if pd.notna(chosen_mape):
            print(f'Точность: {round(100-chosen_mape, 0)}%')

        if chosen_source == "weights":
            method_label = "веса"
            method_detail = f"({float(w1):.2f}, {float(w2):.2f}, {float(w3):.2f})"
        elif chosen_source == "ml":
            method_label = "МЛ"
            method_detail = best_ml_model or ""
        else:
            method_label = "старая"
            method_detail = "старая"

        fc_df = finalize_forecast_output(fc_df, hours_calculator)

        mape_row = pd.DataFrame([{
            "KSSS": k,
            "MAPE_ML": round(mape_ml, 3) if pd.notna(mape_ml) else np.nan,
            "ML_model": best_ml_model or "",
            "MAPE_weights": round(mape_weights, 3) if pd.notna(mape_weights) else np.nan,
            "MAPE_cur": round(mape_cur, 3) if pd.notna(mape_cur) else np.nan,
            "Способ прогноза": method_label,
            "Детали прогноза": method_detail
        }])
        append_to_csv(mape_row, mape_out_csv)

        append_to_csv(fc_df, out_csv)
        forecasted_ksss += 1
        finish_time = time.time()
        print(f'Обработали файл за \033[1m{round(finish_time - start, 2)} сек.\033[0m')
        in_work += (finish_time - start)
        print(f'Накопленное время работы функций: \033[1m{round(in_work, 0)} сек.\033[0m')
    finally:
        write_progress(forecasted_ksss, total_ksss)

if stop_requested():
    print(f'--- Прогноз остановлен пользователем ---\n')
else:
    print(f'--- Программа отработала! ---\n')
    print(f'Файл прогноза: {out_csv}')
    print(f'Файл MAPE: {mape_out_csv}')
    print_mape_stats(mape_out_csv)
    recommendations_path = None
    if not stop_requested():
        recommendations_path = build_recommendations_export(
            out_csv,
            target_month_start,
            target_month_end,
            now,
        )
    if recommendations_path is not None:
        print(f"Файл рассылки рекомендаций: {recommendations_path}")
    print(f'Для выполнения потребовалось {round(time.time() - start_all, 0)} сек.')
